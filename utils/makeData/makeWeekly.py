# -*- coding: utf-8 -*-
import copy
import os

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, colors
from openpyxl.utils import get_column_letter  # 根据列的数字返回字母 get_column_letter(2)  # B
from flask import g

from utils.util.fileUtil import TEMP_FILE_ADDRESS, FileUtil
from utils.makeData.makeExcel import Excel
from utils.util.timeUtil import get_week_start_and_end

# 设置边框线
side = Side(style="thin", color=colors.BLACK)  # 黑色，细线
border = Border(left=side, right=side, top=side, bottom=side)

# 居中格式
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def merge_cells(sheet, start_c, end_c, start_r, end_r):
    """ 合并单元格 """
    sheet.merge_cells(start_column=start_c, end_column=end_c, start_row=start_r, end_row=end_r)


def make_current_weekly_excel(product_dict, weekly_data, user_dict):
    """ 生成当前周的周报 """
    start, end = get_week_start_and_end(1)
    last_start = f'{start.year}' \
                 f'{start.month if start.month > 10 else f"0{start.month}"}' \
                 f'{start.day if start.day > 10 else f"0{start.day}"}'
    last_end = f'{end.year}' \
               f'{end.month if end.month > 10 else f"0{end.month}"}' \
               f'{end.day if end.day > 10 else f"0{end.month}"}'
    file_title = f'测试组周报({last_start}-{last_end})'
    file_name = f'{file_title}_{g.user_id}.xlsx'  # 避免多人导出同一时间段周报时出现死锁问题，每一份文件后都加上用户id
    file_path = os.path.join(TEMP_FILE_ADDRESS, file_name)

    FileUtil.delete_file(file_path)

    excel = Excel(file_path, is_delete_old=True)
    sheet = excel.get_sheet_obj("会议纪要")

    last_week_data, current_week_data = split_week_data(weekly_data)
    build_current_weekly(
        excel, sheet, file_title, product_dict, last_week_data, current_week_data, user_dict
    )

    return file_name


def split_week_data(week_data):
    """ 把当前周和上周的数据分开 """
    last_week_data, current_week_data = [], []
    current_week_start, current_week_end = get_week_start_and_end(0)
    for data in week_data.get("data", []):
        if data["start_time"] < current_week_start:  # 开始时间小于当前周的开始时间，则说明是上周的周报
            last_week_data.append(data)
        else:
            current_week_data.append(data)
    return last_week_data, current_week_data


def build_current_weekly(
        excel, sheet, file_title, product_dict, last_week_data, current_week_data, user_dict):
    """ 生成当前周的周报 """
    build_title(sheet, file_title)  # 表头

    build_class_type(sheet, 2, "一、需要其它小组协助/问题")
    build_detail_title(sheet, 3, ["序号", "项目", "", "问题描述", "责任团队", "备注"])

    build_class_type(sheet, 5, "二、本周提测版本打回情况")
    build_detail_title(sheet, 6, ["序号", "项目", "", "问题描述", "责任团队", "备注"])

    # 上周任务
    build_class_type(sheet, 8, "三、上周测试任务跟踪")
    build_detail_title(sheet, 9, ["产品", "项目", "迭代版本", "任务描述", "测试进展", "测试人员", "备注"], False)
    build_task_item(sheet, 10, product_dict, last_week_data, user_dict)

    # 本周任务
    current_next_item_start = 10 + sum([len(data["task_item"]) for data in last_week_data])
    build_class_type(sheet, current_next_item_start, "四、本周测试任务跟踪")
    build_detail_title(sheet, current_next_item_start + 1, ["产品", "项目", "迭代版本", "任务描述", "测试进展", "测试人员", "备注"], False)
    build_task_item(sheet, current_next_item_start + 2, product_dict, current_week_data, user_dict)

    excel.auto_column_size(sheet)  # 宽度自适应

    # 绘制边框线
    for col in range(sheet.max_column):
        col_name = get_column_letter(col + 1)  # 列名
        for row in range(sheet.max_row):
            sheet[f"{col_name}{str(row + 1)}"].border = border


def build_title(sheet, file_title):
    """ 构建表头 """
    sheet.cell(row=1, column=1, value=file_title).font = Font(size="24", b=True)  # 第一行文字内容和字体大小
    sheet.row_dimensions[1].height = 26  # 设置行高
    sheet["A1"].alignment = center  # 单元格居中格式
    merge_cells(sheet, 1, 7, 1, 1)  # 合并表头单元格
    sheet["A1"].fill = PatternFill(start_color="3366CC", fill_type="solid")  # 绘制背景色


def build_class_type(sheet, row, value):
    """ 问题分类title """
    sheet.cell(row=row, column=1, value=value).font = Font(size="12", b=True)  # 文字内容和字体大小
    merge_cells(sheet, 1, sheet.max_column, row, row)  # 合并单元格
    sheet[f"A{str(row)}"].fill = PatternFill(start_color="DDDDDD", fill_type="solid")  # 绘制背景色


def build_detail_title(sheet, row, title_list, is_merge_last=True):
    """ 详情项的title """
    for index, value in enumerate(title_list):
        # 设置字体
        sheet.cell(row=row, column=index + 1, value=value).font = Font(size="12", b=True)

        # 绘制背景色
        for col in range(sheet.max_column):
            sheet[f"{get_column_letter(col + 1)}{str(row)}"].fill = PatternFill(start_color="DDDDDD", fill_type="solid")

    # 合并最后一个有内容的单元格 + 后面一个单元格
    # if is_merge_last:
    #     sheet.merge_cells(start_column=sheet.max_column, end_column=sheet.max_column + 1, start_row=row, end_row=row)


def parse_task_version(product_dict, week_data):
    """ 解析任务版本号 """
    container = copy.deepcopy(product_dict)
    for product_id, product_detail in container.items():
        for project_id, project_detail in product_detail["project"].items():

            for index, task_item in enumerate(week_data):
                version = task_item["version"].strip()
                if int(task_item.get("product_id")) == product_id:
                    if not task_item.get("project_id"):  # 如果没有所属项目，就把版本放到产品下
                        if version not in product_detail["project"][product_id]["version"]:
                            product_detail["project"][product_id]["version"][version] = {"total": 0, "items": []}
                            product_detail["project"][product_id]["total"] += 1
                        # 把周报数据放到版本下
                        product_detail["project"][product_id]["version"][version]["items"].append(task_item)
                        product_detail["project"][product_id]["version"][version]["total"] += 1
                        week_data.pop(index)
                    elif int(task_item.get("project_id")) == project_id:
                        if version not in project_detail["version"]:
                            project_detail["version"][version] = {"total": 0, "items": []}
                            project_detail["total"] += 1
                        # 把周报数据放到版本下
                        project_detail["version"][version]["items"].append(task_item)
                        project_detail["version"][version]["total"] += 1
                        week_data.pop(index)

    return container


def build_task_item(sheet, start_row, product_dict, week_data, user_dict):
    """ 任务详情 """
    pro_container = parse_task_version(product_dict, copy.deepcopy(week_data))  # 根据版本号归属到项目下
    next_product_start, next_project_start, next_version_start, next_item_start = start_row, start_row, start_row, start_row
    for product_id, product_detail in pro_container.items():
        need_merge_product_title = False
        product_name = product_detail["name"]

        for project_id, project_detail in product_detail["project"].items():
            need_merge_project_title = False
            project_name = project_detail["name"]

            for version, version_detail in project_detail["version"].items():
                need_merge_version_title = False

                for task_data in version_detail["items"]:  # 循环写入周报
                    user_name = user_dict[task_data.get("create_user")]
                    for index, task_item in enumerate(task_data["task_item"]):  # 循环写入周报的详情项
                        for item_index, value in enumerate([
                            product_name, project_name, version, task_item.get("key"), task_item.get("value"),
                            user_name, task_item.get("remark")
                        ]):
                            sheet.cell(row=next_item_start, column=item_index + 1, value=value).font = Font(size="12")
                        need_merge_version_title, need_merge_project_title, need_merge_product_title = True, True, True
                        next_item_start += 1

                # 合并版本号并居中
                if need_merge_version_title:
                    merge_cells(sheet, 3, 3, next_version_start, next_item_start - 1)
                    sheet[f"C{next_version_start}"].alignment = center
                next_version_start = next_item_start  # 下轮版本写入的开始行

            # 合并项目名并居中
            if need_merge_project_title:
                sheet[f"B{next_project_start}"].alignment = center
                merge_cells(sheet, 2, 2, next_project_start, next_version_start - 1)
            next_project_start = next_version_start  # 下轮项目写入的开始行

        # 合并产品名并居中
        if need_merge_product_title:
            merge_cells(sheet, 1, 1, next_product_start, next_project_start - 1)
            sheet[f"A{next_product_start}"].alignment = center
        next_product_start = next_project_start  # 下轮产品写入的开始行


def make_weekly_excel(weekly_data, form, user_dict):
    """ 生成其他时间段的周报 """
    pass
