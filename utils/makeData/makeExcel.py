# -*- coding: utf-8 -*-
import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class Excel:
    """ Excel操作 """

    def __init__(self, file_name, is_delete_old=False):
        self.file_name = file_name
        self.is_delete_old = is_delete_old
        self.excel = self.get_excel()

    def __del__(self):
        """ 保存Excel内容 """
        self.excel.save(self.file_name)

    def get_excel(self):
        """ 获取EXCEL对象，没有则自动创建 """
        if self.is_delete_old and os.path.exists(self.file_name):
            os.remove(self.file_name)
        if not os.path.exists(self.file_name):
            Workbook().save(self.file_name)
        return load_workbook(self.file_name)

    def get_sheet_obj(self, sheet_name):
        """ 获取指定表对象，没有则自动创建 """
        if sheet_name not in self.excel.sheetnames:
            self.excel.create_sheet(sheet_name, 0)
        return self.excel[sheet_name]

    def get_sheet_cols(self, sheet_obj):
        """ 获取表格有数据的列数，取第一行的值出来遍历，最后一个有值的列的索引为最大列数 """
        title_field_list = []
        for index, col in enumerate(list(sheet_obj.rows)[0]):
            if col.value:
                title_field_list.append(col.value)
            else:
                break
        return {"title_list": title_field_list, "total": index}

    def get_sheet_rows(self, sheet_obj):
        """ 获取表格有数据的行数 """
        row_list = []
        for index, row in enumerate(sheet_obj.rows):

            # 第一行数据为表头，不要
            if index == 0:
                continue

            if row[0].value:
                # row_list.append([cell.value for cell in row])
                row_list.append(row)
            else:
                break
        return {"row_list": row_list, "total": index}

    def get_sheet_data(self, sheet_name):
        sheet = self.get_sheet_obj(sheet_name)

        # 列数据
        cols = self.get_sheet_cols(sheet)
        col_list, col_total = cols["title_list"], cols["total"]

        # 行数据
        rows = self.get_sheet_rows(sheet)
        row_list, row_total = rows["row_list"], rows["total"]

        # 压缩并转为字典
        return [dict(zip(col_list, [row[i].value for i in range(col_total)])) for row in row_list]

    def auto_column_size(self, sheet):
        """ 宽度自适应，把每一列中内容最长的长度作为列的宽度 """
        max_row, max_column = sheet.max_row, sheet.max_column
        for col in range(max_column):

            max_widths, col_name = 1, get_column_letter(col + 1)

            # 获取当前列下的所有数据，取单元格值最长的为单元格最大值
            for row in range(max_row):
                content_length = len(sheet[f'{col_name}{row + 1}'].value or "")
                max_widths = content_length if content_length > max_widths else max_widths

            # 把当前列所有单元格的长度设为最大长度
            sheet.column_dimensions[col_name].width = max_widths
